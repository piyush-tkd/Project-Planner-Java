#!/usr/bin/env python3
"""Generate 5 test Excel variations from the base eng_portfolio_planner_clean.xlsx.
Same resources/projects, different sizes/durations/patterns to produce varied utilization."""

import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

BASE = "/Users/piyushbaheti/HTML - Project Planner/eng_portfolio_planner_clean.xlsx"
OUT_DIR = "/Users/piyushbaheti/HTML - Project Planner"

# POD Planning sheet structure: row 4+ data, cols: Project(A), POD(B), Size(C), Complexity(D), Pattern(E), Start(F), Duration(G)
# Projects sheet: row 4+ data, cols: Project(A), Priority(B), Owner(C), Start(D), End(E), Duration(F), Pattern(G), Status(H), Target(I), Notes(J)

SIZES_ORDERED = ['XS', 'S', 'S+', 'M', 'M+', 'L', 'L+', 'XL', 'XXL', 'XXXL', 'Mega', 'Program']
SIZE_IDX = {s: i for i, s in enumerate(SIZES_ORDERED)}

def shift_size(size, delta):
    if size not in SIZE_IDX:
        return size
    new_idx = max(0, min(len(SIZES_ORDERED) - 1, SIZE_IDX[size] + delta))
    return SIZES_ORDERED[new_idx]

def generate_variant(variant_num, description, modify_fn):
    wb = load_workbook(BASE)
    modify_fn(wb)
    out_path = os.path.join(OUT_DIR, f"test_variant_{variant_num}_{description}.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")
    return out_path

# ═══════════════════════════════════════════════
# Variant 1: Balanced — all pods green/yellow
# Strategy: Downsize everything to XS/S, short durations
# ═══════════════════════════════════════════════
def variant_balanced(wb):
    pod_sheet = wb['POD Planning']
    proj_sheet = wb['Projects']
    # Downsize all POD planning entries by 3 levels
    for row in pod_sheet.iter_rows(min_row=4, max_col=7):
        if row[0].value and row[2].value:
            row[2].value = shift_size(str(row[2].value), -3)
    # Shorten all projects to 2-3 months
    for row in proj_sheet.iter_rows(min_row=4, max_col=10):
        if row[0].value and row[3].value:
            start = row[3].value  # e.g. "M1"
            if isinstance(start, str) and start.startswith('M'):
                sm = int(start[1:])
                dur = 2
                em = min(sm + dur - 1, 12)
                row[4].value = f"M{em}"
                row[5].value = dur
                row[6].value = 'Flat'

# ═══════════════════════════════════════════════
# Variant 2: All pods critical/red
# Strategy: Upsize everything by 2 levels, long durations
# ═══════════════════════════════════════════════
def variant_all_red(wb):
    pod_sheet = wb['POD Planning']
    proj_sheet = wb['Projects']
    # Upsize all POD planning entries by 2 levels
    for row in pod_sheet.iter_rows(min_row=4, max_col=7):
        if row[0].value and row[2].value:
            row[2].value = shift_size(str(row[2].value), 2)
    # Extend all projects to 10-12 months, start early
    for row in proj_sheet.iter_rows(min_row=4, max_col=10):
        if row[0].value and row[3].value:
            row[3].value = 'M1'
            row[4].value = 'M12'
            row[5].value = 12
            row[6].value = 'Flat'

# ═══════════════════════════════════════════════
# Variant 3: Mixed — Portal pods red, others green
# Strategy: Portal V1/V2 get XL sizes, others get XS/S
# ═══════════════════════════════════════════════
def variant_mixed(wb):
    pod_sheet = wb['POD Planning']
    proj_sheet = wb['Projects']
    red_pods = {'Portal V1', 'Portal V2'}
    for row in pod_sheet.iter_rows(min_row=4, max_col=7):
        if row[0].value and row[1].value and row[2].value:
            pod = str(row[1].value).strip()
            if pod in red_pods:
                row[2].value = shift_size(str(row[2].value), 2)
            else:
                row[2].value = shift_size(str(row[2].value), -3)
    # Shorten non-portal projects, extend portal ones
    portal_projects = set()
    for row in pod_sheet.iter_rows(min_row=4, max_col=7):
        if row[0].value and row[1].value:
            pod = str(row[1].value).strip()
            if pod in red_pods:
                portal_projects.add(str(row[0].value).strip())
    for row in proj_sheet.iter_rows(min_row=4, max_col=10):
        if row[0].value and row[3].value:
            name = str(row[0].value).strip()
            if name in portal_projects:
                row[3].value = 'M1'
                row[4].value = 'M12'
                row[5].value = 12
            else:
                start = row[3].value
                if isinstance(start, str) and start.startswith('M'):
                    sm = int(start[1:])
                    dur = 2
                    em = min(sm + dur - 1, 12)
                    row[4].value = f"M{em}"
                    row[5].value = dur

# ═══════════════════════════════════════════════
# Variant 4: Seasonal crunch — big spike in M5-M8
# Strategy: Many projects compressed into mid-year
# ═══════════════════════════════════════════════
def variant_seasonal(wb):
    pod_sheet = wb['POD Planning']
    proj_sheet = wb['Projects']
    # Keep sizes as-is but compress all projects to M5-M8
    for i, row in enumerate(proj_sheet.iter_rows(min_row=4, max_col=10)):
        if row[0].value and row[3].value:
            # Alternate: half start M5, half start M6
            if i % 2 == 0:
                row[3].value = 'M5'
                row[4].value = 'M8'
                row[5].value = 4
            else:
                row[3].value = 'M6'
                row[4].value = 'M9'
                row[5].value = 4
            row[6].value = 'Flat'

# ═══════════════════════════════════════════════
# Variant 5: One overloaded pod (Integrations), rest fine
# Strategy: Make Integrations XL/XXL, downsize everything else
# ═══════════════════════════════════════════════
def variant_one_hot(wb):
    pod_sheet = wb['POD Planning']
    proj_sheet = wb['Projects']
    hot_pod = 'Integrations'
    for row in pod_sheet.iter_rows(min_row=4, max_col=7):
        if row[0].value and row[1].value and row[2].value:
            pod = str(row[1].value).strip()
            if pod == hot_pod:
                row[2].value = shift_size(str(row[2].value), 3)
            else:
                row[2].value = shift_size(str(row[2].value), -3)
    # Extend Integration-heavy projects, shorten others
    integ_projects = set()
    for row in pod_sheet.iter_rows(min_row=4, max_col=7):
        if row[0].value and row[1].value:
            pod = str(row[1].value).strip()
            if pod == hot_pod:
                integ_projects.add(str(row[0].value).strip())
    for row in proj_sheet.iter_rows(min_row=4, max_col=10):
        if row[0].value and row[3].value:
            name = str(row[0].value).strip()
            if name in integ_projects:
                row[3].value = 'M1'
                row[4].value = 'M12'
                row[5].value = 12
                row[6].value = 'Ramp Up'
            else:
                start = row[3].value
                if isinstance(start, str) and start.startswith('M'):
                    sm = int(start[1:])
                    dur = 2
                    em = min(sm + dur - 1, 12)
                    row[4].value = f"M{em}"
                    row[5].value = dur

# Generate all 5 variants
generate_variant(1, "balanced_all_green", variant_balanced)
generate_variant(2, "all_pods_red", variant_all_red)
generate_variant(3, "portal_pods_red_others_green", variant_mixed)
generate_variant(4, "seasonal_midyear_crunch", variant_seasonal)
generate_variant(5, "integrations_overloaded", variant_one_hot)

print("\nDone! All 5 test Excel files generated.")
